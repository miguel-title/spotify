import requests
import argparse
import json
import csv
import unicodedata

import os
import sys

import openpyxl

if sys.version_info >= (3, 0):
    import configparser as ConfigParser
else:
    import ConfigParser

import time

class spotifyApp():
	def __init__(self):
		self.accessToken = ""
		self.client_id = '8a14ebc91cf34265b1acbfb777b946ab'
		self.client_secret = 'fe7ad45553294aff829b4a101f059fb5'
		self.scope = 'playlist-modify-public playlist-modify-private playlist-read-collaborative'
		self.refresh_token =  'AQAsQJRqXNVHkMsfAtvKciUg8LuinmLEi92NnDXDk2ePwsIJMc14Qci6Lhk7F-oAXOkOIcTnSJOOdMumR4yOMLVX8IbmRCzxIKZJ7jgytLwxAaEfYcY6pZMURNfCINJfwxs'
		self.excelheader = [
							'Year',
							'Artist',
							'Album',
							'Url',
							'Popularity'
							]

	def get_access_token(self):
		url = 'https://accounts.spotify.com/api/token'
		payload = {
		'grant_type': 'refresh_token',
		'refresh_token': self.refresh_token
		}
		auth = (self.client_id, self.client_secret)
		token = requests.post(url, data=payload, auth=auth).json()
		self.accessToken = token['access_token']
		return

	def getData(self, startyear, endyear, quantify, updatepopularity):
		# csvfile = open(outputpath, 'w', newline='\n')#, encoding="utf-8-sig")#utf-8-sig, utf8
		# writer = csv.DictWriter(csvfile, delimiter=",", fieldnames=self.excelheader)
		# writer.writeheader()

		for year in range(int(startyear), int(endyear) + 1, 1):

			outputfile = "output_{}.xlsx".format(year)
			isFileExist = False
			if os.path.isfile(outputfile):
				print ("File exist")
				isFileExist = True
			else:
				print ("File not exist")

			oldalbumurldata = []

			if isFileExist == True:
				wb = openpyxl.load_workbook(outputfile)
				sheet = wb.active
				grow = sheet.max_row + 1
				for row in range(grow):
					if row < 2:
						continue
					oldalbumurldata.append(sheet.cell(row = row, column = 4).value)

				if updatepopularity == "yes":
					print("update the popularity")
					self.get_access_token()
					oldrow = 2
					for oldalbumurl in oldalbumurldata:
						while True:
							try:
								albumid = oldalbumurl.replace('https://open.spotify.com/album/', '')
								headers = {'Origin': 'https://open.spotify.com',
									'Accept-Encoding': 'gzip, deflate, br',
									'Accept-Language': 'en',
									'Authorization': 'Bearer ' + self.accessToken,
									'Accept': 'application/json',
									# 'Referer': 'https://open.spotify.com/search/albums/year^%^3A1980',
									'Authority': 'api.spotify.com',
									'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36'}

								presponse = requests.get('https://api.spotify.com/v1/albums/{}'.format(albumid), headers=headers).json()

								try:
									varpopularity = presponse['popularity']
								except:
									varpopularity = 0

								print("------update row : {}".format(oldrow))
								print('https://api.spotify.com/v1/albums/{}'.format(albumid))
								sheet.cell(row = oldrow, column=5).value = varpopularity
								oldrow += 1
							except:
								time.sleep(1)
								continue
							break

			else:
				wb = openpyxl.Workbook() 
				sheet = wb.active 
				for i in range(5):
					sheet.cell(row = 1, column=i + 1).value = self.excelheader[i]
				grow = 2

			for i in range(0, int(quantify), 50):
				self.get_access_token()
				print('-------------{}'.format(i))
				headers = {'Origin': 'https://open.spotify.com',
							'Accept-Encoding': 'gzip, deflate, br',
							'Accept-Language': 'en',
							'Authorization': 'Bearer ' + self.accessToken,
							'Accept': 'application/json',
							# 'Referer': 'https://open.spotify.com/search/albums/year^%^3A1980',
							'Authority': 'api.spotify.com',
							'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36'}

				response = requests.get('https://api.spotify.com/v1/search?query=year%3A{}&type=album&include_external=audio&market=US&offset={}&limit=50'.format(year, i), headers=headers).json()
				
				try:
					data = response['albums']
				except:
					print(response)
					continue

				for album in response['albums']['items']:

					output = {
							'Year':'',
							'Artist':'',
							'Album':'',
							'Url':'',
							'Popularity':''
					}

					varyear = year
					try:
						varartist = album['artists'][0]['name'].encode('utf-8').strip()
					except:
						varartist = ""

					try:
						varalbum = album['name'].encode('utf-8').strip()
					except:
						varalbum = ""

					try:
						varurl = album['external_urls']['spotify'].encode('utf-8').strip()
					except:
						varurl = ""

					try:
						varhref = album['href']

						presponse = requests.get(varhref, headers=headers).json()
						
						try:
							varpopularity = presponse['popularity']
						except:
							varpopularity = 0
					except:
						varpopularity = 0

					print("Year:{}, Artist:{}, Album:{}, Url:{}, Popularity:{}".format(varyear, varartist, varalbum, varurl,varpopularity))

					# write the data to excel
					output['Year'] = varyear
					output['Artist'] = varartist
					output['Album'] = varalbum
					output['Url'] = varurl
					output['Popularity'] = varpopularity

					#writer.writerow(output)

					if isFileExist == True:
						if output['Url'].decode('utf-8') in oldalbumurldata:
							if updatepopularity == 'yes':
								sheet.cell(row = oldalbumurldata.index(output['Url'].decode('utf-8')) + 2, column=5).value = output['Popularity']
							continue
						else:
							sheet.cell(row = grow, column=1).value = output['Year']
							sheet.cell(row = grow, column=2).value = output['Artist']
							sheet.cell(row = grow, column=3).value = output['Album']
							sheet.cell(row = grow, column=4).value = output['Url']
							sheet.cell(row = grow, column=5).value = output['Popularity']
							grow += 1
					else:
						sheet.cell(row = grow, column=1).value = output['Year']
						sheet.cell(row = grow, column=2).value = output['Artist']
						sheet.cell(row = grow, column=3).value = output['Album']
						sheet.cell(row = grow, column=4).value = output['Url']
						sheet.cell(row = grow, column=5).value = output['Popularity']
						grow += 1
		#csvfile.close()

			wb.save(outputfile)

def load_config():
	defaults = {
		'output': '',
		'startyear': '',
		'endyear': '',
		'quantify': '',
	}
	_settings_dir = "./"
	config_file = os.path.join(_settings_dir, "config.ini")
	if os.path.exists(config_file):
		print('Existing config.ini')
		try:
		# config = ConfigParser.SafeConfigParser()
			config = ConfigParser.ConfigParser()
			config.read(config_file)
			if config.has_section("global"):
				config_items = dict(config.items("global"))

				defaults['updatepopularity'] = config_items['updatepopularity']
				defaults['startyear'] = config_items['startyear']
				defaults['endyear'] = config_items['endyear']
				defaults['quantify'] = config_items['quantify']
		except ConfigParser.Error as e:
			print("\nError parsing config file: " + config_file)
			print(str(e))
			exit(1)

	return defaults

def startProcess():
	# parser = argparse.ArgumentParser()
	# parser.add_argument("-y", "--year", default=1980, type=int)
	# parser.add_argument("-q", "--quantify", default=100, type=int)
	# parser.add_argument("-o", "--outputpath", default='output.csv', type=str)
	# args = parser.parse_args()

	config_option = load_config()

	updatepopularity = config_option['updatepopularity']
	quantify = config_option['quantify']
	startyear = config_option['startyear']
	endyear = config_option['endyear']

	app = spotifyApp()
	app.getData(startyear, endyear, quantify, updatepopularity)

if __name__=="__main__":
	startProcess()
	print("------------finish------------")